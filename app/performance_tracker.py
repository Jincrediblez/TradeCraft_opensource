#!/usr/bin/env python3
"""Performance tracker: reads net-worth xlsx and computes monthly returns."""

import json
import os
from pathlib import Path
from typing import Dict, List, Optional

from openpyxl import load_workbook


try:  # normal package import (server, cli, pytest)
    from app.state_store import read_json as _store_read_json, atomic_write_text as _atomic_write
except ImportError:  # standalone script execution puts app/ on sys.path
    from state_store import read_json as _store_read_json, atomic_write_text as _atomic_write

ROOT = Path(__file__).resolve().parent.parent
STATE_DIR = ROOT / "data" / "state"
LOCAL_XLSX_PATH = ROOT / "data" / "inbox" / "收益曲线.xlsx"
CACHE_FILE = STATE_DIR / "performance.json"

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]
MONTH_MAP = {m: i + 1 for i, m in enumerate(MONTHS)}


def _candidate_xlsx_paths() -> List[Path]:
    configured = str(os.environ.get("TRADECRAFT_PERFORMANCE_FILE", "")).strip()
    candidates = [Path(configured).expanduser()] if configured else []
    candidates.append(LOCAL_XLSX_PATH)
    return candidates


def _load_cached_payload() -> Optional[dict]:
    cached = _store_read_json(CACHE_FILE, None)
    if isinstance(cached, dict) and cached.get("series"):
        return cached
    return None


def _read_xlsx() -> Optional[dict]:
    """Parse the xlsx and return structured performance data."""
    errors = []
    xlsx_path = None
    wb = None
    for path in _candidate_xlsx_paths():
        try:
            if not path.exists():
                continue
            wb = load_workbook(path, data_only=True)
            xlsx_path = path
            break
        except PermissionError as exc:
            errors.append(f"{path}: {exc}")
        except OSError as exc:
            errors.append(f"{path}: {exc}")

    if wb is None:
        if errors:
            raise PermissionError("; ".join(errors))
        return None

    ws = wb.active

    # Header row
    headers = []
    col = 2
    while True:
        val = ws.cell(row=1, column=col).value
        if not val:
            break
        headers.append(str(val).replace("FY", ""))  # "2018FY" -> "2018"
        col += 1

    series: List[dict] = []
    years = []
    current_year = None

    for year_idx, year in enumerate(headers):
        years.append(year)
        year_values = []
        for month_name in MONTHS:
            row = MONTH_MAP[month_name] + 1  # row 2 = Jan
            val = ws.cell(row=row, column=year_idx + 2).value
            if val is None:
                break
            try:
                amount = float(val)
            except (TypeError, ValueError):
                break
            year_values.append(amount)

        if not year_values:
            continue

        # Determine current year (latest with data)
        if current_year is None or year_values:
            current_year = year

        prev_value = None
        prev_year_last_value = series[-1]["value"] if series else None
        ytd_base_value = prev_year_last_value if prev_year_last_value not in (None, 0) else (year_values[0] if year_values else 0)
        ytd_base_month = series[-1]["month"] if prev_year_last_value not in (None, 0) and series else None
        ytd_base_year = series[-1]["year"] if prev_year_last_value not in (None, 0) and series else None

        for i, amount in enumerate(year_values):
            month = MONTHS[i]
            time_key = f"{year}-{i+1:02d}-01"

            # Monthly return
            if prev_value is not None and prev_value != 0:
                monthly_return = (amount - prev_value) / prev_value * 100
            else:
                monthly_return = 0.0

            # YTD return is measured from the previous year's final value when available.
            if ytd_base_value != 0:
                ytd_return = (amount - ytd_base_value) / ytd_base_value * 100
            else:
                ytd_return = 0.0

            series.append({
                "time": time_key,
                "year": year,
                "month": month,
                "monthNum": i + 1,
                "value": round(amount, 2),
                "valueWan": round(amount / 10000, 1),
                "monthlyReturn": round(monthly_return, 2),
                "ytdReturn": round(ytd_return, 2),
                "ytdBaseValue": round(ytd_base_value, 2) if ytd_base_value else 0.0,
                "ytdBaseValueWan": round(ytd_base_value / 10000, 1) if ytd_base_value else 0.0,
                "ytdBaseYear": ytd_base_year,
                "ytdBaseMonth": ytd_base_month,
            })

            prev_value = amount

    if not series:
        return None

    # Compute cumulative return relative to the very first data point
    first_value = series[0]["value"]
    for item in series:
        if first_value != 0:
            item["cumulativeReturn"] = round((item["value"] - first_value) / first_value * 100, 2)
        else:
            item["cumulativeReturn"] = 0.0

    latest = series[-1]
    summary = {
        "startValue": series[0]["value"],
        "startValueWan": series[0]["valueWan"],
        "latestValue": latest["value"],
        "latestValueWan": latest["valueWan"],
        "totalReturn": latest["cumulativeReturn"],
        "latestYearYtd": latest["ytdReturn"],
        "latestYear": latest["year"],
        "latestMonth": latest["month"],
    }

    return {
        "years": years,
        "currentYear": current_year,
        "series": series,
        "summary": summary,
        "sourcePath": (
            str(xlsx_path.relative_to(ROOT))
            if xlsx_path and xlsx_path.is_relative_to(ROOT)
            else ("configured performance workbook" if xlsx_path else "")
        ),
    }


def build_performance_payload(force_refresh: bool = False) -> dict:
    """Return performance data, using cache when available."""
    if not force_refresh:
        cached = _load_cached_payload()
        if cached:
            return cached

    try:
        data = _read_xlsx()
    except PermissionError as exc:
        cached = _load_cached_payload()
        if cached:
            cached = dict(cached)
            cached["stale"] = True
            cached["warning"] = (
                "无法读取配置的收益曲线工作簿，当前继续使用业绩缓存。"
                "如需刷新，请把文件放到 data/inbox/收益曲线.xlsx，或给本地服务授予文件访问权限。"
            )
            cached["sourceError"] = str(exc)
            return cached
        return {
            "error": (
                "无法读取收益曲线.xlsx。请把文件放到 data/inbox/收益曲线.xlsx，"
                "或检查 TRADECRAFT_PERFORMANCE_FILE 的访问权限。"
            ),
            "sourceError": str(exc),
        }

    if data is None:
        cached = _load_cached_payload()
        if cached:
            cached = dict(cached)
            cached["stale"] = True
            cached["warning"] = "未找到收益曲线.xlsx，当前继续使用业绩缓存。"
            return cached
        return {"error": "XLSX not found or empty"}

    data["stale"] = False
    data["warning"] = ""
    _atomic_write(CACHE_FILE, json.dumps(data, ensure_ascii=False, indent=2))
    return data


def refresh_performance() -> dict:
    """Force re-read xlsx and update cache."""
    return build_performance_payload(force_refresh=True)


if __name__ == "__main__":
    result = build_performance_payload(force_refresh=True)
    print(json.dumps(result, ensure_ascii=False, indent=2))
